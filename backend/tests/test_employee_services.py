from io import BytesIO

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.models import (
    Employee,
    EmploymentStatus,
    Evaluation,
    EvaluationPeriod,
    PeriodSeason,
    PeriodStatus,
    SubmissionStatus,
    User,
    UserRole,
)
from app.services.employee_import import import_employees_from_excel
from app.services.employee_lifecycle import retire_employee
from app.services.excel_export import build_period_export
from app.main import load_assessment_template, load_self_template


def fake_hash(password: str) -> str:
    return f"test-hash-{password}"


@pytest.fixture()
def db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()

    admin_user = User(
        employee_id="admin1",
        password_hash=fake_hash("testpass123"),
        role=UserRole.ADMIN,
    )
    session.add(admin_user)
    session.flush()

    ev1 = Employee(
        employee_id="ev1",
        name="評価者1",
        assignment="施設",
        job_type="介護",
        employment_status=EmploymentStatus.ACTIVE,
        user_id=admin_user.id,
    )
    ev2 = Employee(
        employee_id="ev2",
        name="評価者2",
        assignment="施設",
        job_type="介護",
        employment_status=EmploymentStatus.ACTIVE,
    )
    session.add_all([ev1, ev2])
    session.flush()

    period = EvaluationPeriod(
        name="テスト期間",
        season=PeriodSeason.SUMMER,
        fiscal_year=8,
        status=PeriodStatus.ACTIVE,
    )
    session.add(period)
    session.commit()
    yield session
    session.close()


def _build_import_xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "社員ID",
            "氏名",
            "配属",
            "職種",
            "勤続年数",
            "評価者1社員ID",
            "評価者2社員ID",
        ]
    )
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_import_creates_evaluations_only_for_imported_rows(db):
    existing = Employee(
        employee_id="old1",
        name="既存職員",
        assignment="施設",
        job_type="介護",
        employment_status=EmploymentStatus.ACTIVE,
    )
    db.add(existing)
    db.commit()

    admin = db.query(User).filter(User.role == UserRole.ADMIN).first()
    period = db.query(EvaluationPeriod).first()
    content = _build_import_xlsx([["new1", "新規", "施設", "介護", 1, "ev1", "ev2"]])

    result = import_employees_from_excel(
        db=db,
        content=content,
        admin_user=admin,
        password_hash_fn=fake_hash,
        default_password="changeme123",
        active_period=period,
    )

    assert result.evaluations_created == 1
    eval_count = db.query(Evaluation).filter(Evaluation.period_id == period.id).count()
    assert eval_count == 1


def test_retire_clears_evaluator_references_and_creates_archive(db, tmp_path, monkeypatch):
    monkeypatch.setattr(
        "app.services.employee_lifecycle.settings.retired_archives_dir",
        str(tmp_path),
    )

    admin = db.query(User).filter(User.role == UserRole.ADMIN).first()
    ev1 = db.query(Employee).filter(Employee.employee_id == "ev1").first()

    subordinate = Employee(
        employee_id="sub1",
        name="部下",
        assignment="施設",
        job_type="介護",
        employment_status=EmploymentStatus.ACTIVE,
        evaluator1_id=ev1.id,
    )
    db.add(subordinate)
    db.commit()

    result = retire_employee(db, ev1, admin, reason="テスト")
    db.commit()

    db.refresh(subordinate)
    assert subordinate.evaluator1_id is None
    assert result.archive_id
    assert list(tmp_path.glob("*.json"))


def test_period_export_generates_xlsx(db):
    period = db.query(EvaluationPeriod).first()
    employee = Employee(
        employee_id="exp1",
        name="出力テスト",
        assignment="施設",
        job_type="介護",
        employment_status=EmploymentStatus.ACTIVE,
    )
    db.add(employee)
    db.flush()
    db.add(Evaluation(period_id=period.id, employee_id=employee.id))
    db.commit()

    data, filename = build_period_export(
        db,
        period,
        load_self_template(),
        load_assessment_template(),
    )
    assert filename.endswith(".xlsx")
    assert data[:2] == b"PK"


def test_bonus_job_title_mapping():
    from app.models import Employee
    from app.services.bonus_workbook import bonus_job_title_label, format_bonus_name

    assert format_bonus_name("宮嶋 くみ子") == "宮嶋\u3000くみ子"
    assert bonus_job_title_label(
        Employee(
            employee_id="x",
            name="加納",
            assignment="",
            job_type="介護支援専門員",
            job_title="リーダー",
        )
    ) == "生活相談員"
    assert bonus_job_title_label(
        Employee(
            employee_id="y",
            name="宮嶋",
            assignment="",
            job_type="看護師",
            job_title="一般",
        )
    ) == "看護職員"


def test_bonus_reflect_name_normalization():
    from app.services.bonus_workbook import normalize_person_name

    assert normalize_person_name("渡邉 昌代") == normalize_person_name("渡邊　昌代")
    assert normalize_person_name("宮嶋 くみ子") == normalize_person_name("宮嶋　くみ子")


def test_bonus_rank_grade_distribution():
    from app.services.bonus_workbook import (
        distribute_grade_counts,
        grade_for_rank,
    )

    rules = [("A", 0.10), ("B", 0.20), ("C", 0.40), ("D", 0.20), ("E", 0.10)]
    assert distribute_grade_counts(10, rules) == [
        ("A", 1),
        ("B", 2),
        ("C", 4),
        ("D", 2),
        ("E", 1),
    ]
    assert grade_for_rank(1, 10, rules) == "A"
    assert grade_for_rank(4, 10, rules) == "C"
    assert grade_for_rank(10, 10, rules) == "E"


def test_bonus_low_score_and_salary_rules():
    from app.services.bonus_workbook import (
        _count_low_other_scores,
        _count_low_self_scores,
        _salary_raise_mark,
    )

    scores = {"item_01": 8, "item_02": 6, "item_03": 10, "item_04": 9, "item_05": 7}
    assert _count_low_self_scores(scores) == 2
    assert _count_low_other_scores(scores) == 1
    assert _salary_raise_mark(scores) == "○"


def test_bonus_evaluation_insights():
    from app.models import Employee
    from app.services.bonus_workbook import build_evaluation_insights

    employee = Employee(
        employee_id="lead1",
        name="テスト",
        assignment="",
        job_type="介護",
        job_title="リーダー",
    )
    evaluation = Evaluation(
        period_id=__import__("uuid").uuid4(),
        employee_id=employee.id,
        self_eval_status=SubmissionStatus.SUBMITTED,
        eval1_status=SubmissionStatus.SUBMITTED,
        eval2_status=SubmissionStatus.SUBMITTED,
        self_eval_data={
            "scores": {
                "item_01": 7,
                "item_02": 8,
                "item_03": 8,
                "item_04": 8,
                "item_05": 8,
                "item_06": 7,
                "item_07": 8,
                "item_08": 8,
                "item_09": 8,
                "item_10": 8,
            }
        },
        eval1_data={"scores": {}, "text_fields": {}},
        eval2_data={
            "scores": {
                "item_01": 6,
                "item_02": 8,
                "item_03": 8,
                "item_04": 8,
                "item_05": 8,
                "item_06": 8,
                "item_07": 8,
                "item_08": 7,
                "item_09": 8,
                "item_10": 8,
            },
            "text_fields": {},
        },
    )
    insights = build_evaluation_insights(employee, evaluation)
    assert insights["cut_self_items"] == "①（1件）"
    assert insights["cut_other_items"] == "①（1件）"
    assert insights["promotion_reference"] == "自己⑥ / 他者⑧"
    assert insights["is_role_holder"] is True


def test_facility_registry_and_evaluator_scope():
    from app.models import Employee, User, UserRole
    from app.services.facilities import (
        DEFAULT_BONUS_FACILITY_KEY,
        build_bonus_preset,
        list_all_facilities,
        resolve_bonus_facility_key,
        resolve_facility_for_assignment,
        validate_evaluator_facility,
    )

    facilities = list_all_facilities()
    assert any(facility.key == "inaha" for facility in facilities)
    assert any(facility.key == "sohara" for facility in facilities)
    preset = build_bonus_preset("inaha")
    assert preset["data_sheet"].startswith("R7")
    assert preset["assignment_contains"] == "サフィールいなは"

    employee = Employee(
        employee_id="e1",
        name="A",
        assignment="サフィールいなは",
        job_type="介護",
        job_title="一般",
    )
    same = Employee(
        employee_id="l1",
        name="L",
        assignment="サフィールいなは",
        job_type="介護",
        job_title="リーダー",
    )
    other = Employee(
        employee_id="l2",
        name="L2",
        assignment="サフィールそはら",
        job_type="介護",
        job_title="リーダー",
    )
    assert resolve_facility_for_assignment(employee.assignment).key == "inaha"
    assert validate_evaluator_facility(employee, same, role_label="第1評価者") is None
    assert validate_evaluator_facility(employee, other, role_label="第1評価者")

    director = Employee(
        employee_id="dir1",
        name="施設長",
        assignment="サフィールそはら",
        job_type="管理",
        job_title="施設長",
    )
    user = User(employee_id="dir1", password_hash="x", role=UserRole.EVALUATOR2, is_admin=True)
    assert resolve_bonus_facility_key(user, director, "sohara") == "sohara"
    try:
        resolve_bonus_facility_key(user, director, "unuma")
        assert False, "expected ValueError"
    except ValueError as exc:
        assert "アクセス" in str(exc)


def test_bonus_reflect_writes_scores(db, tmp_path, monkeypatch):
    from openpyxl import Workbook

    from app.services.bonus_workbook import reflect_evaluations_to_bonus_workbook

    template = tmp_path / "bonus.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "R7\u3000いなは資料 "
    ws["B4"] = "テスト　太郎"
    wb.save(template)

    monkeypatch.setattr(
        "app.services.bonus_workbook.settings.bonus_workbook_template_path",
        str(template),
    )
    monkeypatch.setattr("app.services.bonus_workbook.settings.bonus_workbook_password", "")

    period = db.query(EvaluationPeriod).first()
    employee = Employee(
        employee_id="bonus1",
        name="テスト 太郎",
        assignment="サフィールいなは",
        job_type="介護",
        job_title="一般",
        employment_status=EmploymentStatus.ACTIVE,
    )
    db.add(employee)
    db.flush()
    db.add(
        Evaluation(
            period_id=period.id,
            employee_id=employee.id,
            self_eval_status=SubmissionStatus.SUBMITTED,
            eval1_status=SubmissionStatus.SUBMITTED,
            eval2_status=SubmissionStatus.SUBMITTED,
            self_eval_data={
                "scores": {"item_01": 10, "item_02": 8, "item_03": 8, "item_04": 8, "item_05": 8},
            },
            eval1_data={
                "scores": {
                    "item_01": 8,
                    "item_02": 8,
                    "item_03": 8,
                    "item_04": 8,
                    "item_05": 8,
                },
                "text_fields": {},
            },
                eval2_data={
                    "scores": {
                        "item_01": 8,
                        "item_02": 6,
                        "item_03": 10,
                        "item_04": 8,
                        "item_05": 8,
                    },
                "text_fields": {"evaluator2_note": "所見テスト"},
            },
        )
    )
    db.commit()

    data, result = reflect_evaluations_to_bonus_workbook(db, period, facility_key="inaha")
    assert result.updated_rows == 1
    assert data and data[:2] == b"PK"

    from io import BytesIO

    out_wb = __import__("openpyxl").load_workbook(BytesIO(data), data_only=False)
    out_ws = out_wb["R7\u3000いなは資料 "]
    assert out_ws["D4"].value == 42
    assert out_ws["E4"].value == 40
    assert out_ws["F4"].value == 40
    assert out_ws["I4"].value == 1
    assert out_ws["J4"].value == "○"
    assert out_ws["M4"].value == "所見テスト"


def test_bonus_reflect_writes_facility_director_scores(db, tmp_path, monkeypatch):
    from io import BytesIO

    from openpyxl import Workbook

    from app.services.bonus_workbook import load_bonus_workbook_rows, reflect_evaluations_to_bonus_workbook
    from app.services.facilities import FACILITY_DIRECTORS_BONUS_KEY

    template = tmp_path / "bonus_director.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "R7\u3000いなは資料 "
    ws["B4"] = "施設長　テスト"
    wb.save(template)

    monkeypatch.setattr(
        "app.services.bonus_workbook.settings.bonus_workbook_template_path",
        str(template),
    )
    monkeypatch.setattr("app.services.bonus_workbook.settings.bonus_workbook_password", "")

    period = db.query(EvaluationPeriod).first()
    director = Employee(
        employee_id="dir_bonus",
        name="施設長 テスト",
        assignment="サフィールいなは",
        job_type="管理職",
        job_title="施設長",
        employment_status=EmploymentStatus.ACTIVE,
    )
    db.add(director)
    db.flush()
    db.add(
        Evaluation(
            period_id=period.id,
            employee_id=director.id,
            self_eval_status=SubmissionStatus.SUBMITTED,
            eval1_status=SubmissionStatus.PENDING,
            eval2_status=SubmissionStatus.SUBMITTED,
            self_eval_data={
                "scores": {"item_01": 9, "item_02": 9, "item_03": 9, "item_04": 9, "item_05": 9},
            },
            eval2_data={
                "scores": {
                    "item_01": 8,
                    "item_02": 8,
                    "item_03": 8,
                    "item_04": 8,
                    "item_05": 8,
                },
                "text_fields": {"evaluator2_note": "本部所見"},
            },
        )
    )
    db.commit()

    _, result = reflect_evaluations_to_bonus_workbook(db, period, facility_key="inaha")
    assert result.updated_rows == 1

    out_wb = __import__("openpyxl").load_workbook(template, data_only=True)
    out_ws = out_wb["R7\u3000いなは資料 "]
    assert out_ws["D4"].value == 45
    assert out_ws["E4"].value is None
    assert out_ws["F4"].value == 40

    rows, label = load_bonus_workbook_rows(db, facility_key=FACILITY_DIRECTORS_BONUS_KEY)
    assert label == "施設長"
    assert len(rows) == 1
    assert rows[0]["employee_id"] == "dir_bonus"
    assert rows[0]["self_score"] == 45
    assert rows[0]["eval1_score"] is None
    assert rows[0]["eval2_score"] == 40
    assert rows[0]["facility_label"] == "サフィールいなは"
    assert rows[0]["cut_self_items"] is None
    assert rows[0]["cut_other_items"] is None
    assert rows[0]["promotion_reference"] is None

    assert out_ws["H4"].value is None
    assert out_ws["I4"].value is None
    assert out_ws["J4"].value is None

    _, directors_result = reflect_evaluations_to_bonus_workbook(
        db, period, facility_key=FACILITY_DIRECTORS_BONUS_KEY
    )
    assert directors_result.updated_rows == 1
