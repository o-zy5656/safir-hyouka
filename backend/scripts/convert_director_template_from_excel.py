"""施設長評価シート Excel から JSON テンプレートを生成。

使い方:
  cd backend
  python -m scripts.convert_director_template_from_excel
  python -m scripts.convert_director_template_from_excel /path/to/施設長　評価シート.xlsx
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.database import SessionLocal
from app.main import migrate_schema
from app.models import EvaluationPeriod, FormTemplate, PeriodStatus
from app.services.template_validator import validate_template

DEFAULT_EXCEL = Path(
    "/Users/kawamuraakihiko/Library/CloudStorage/Dropbox/サフィール関係/"
    "賞与・昇給・給与/令和7年度/【いなは】施設長　評価シート.xlsx"
)
TEMPLATES_DIR = Path(__file__).resolve().parent.parent / "templates"

CRITERIA = [
    {"score": 1, "text": "出来ていない"},
    {"score": 2, "text": "だいたい出来ている（５割以上）"},
    {"score": 3, "text": "ほぼ出来ている（８割以上）"},
]


def parse_items(excel_path: Path) -> list[dict]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    items: list[dict] = []
    for row in ws.iter_rows(min_row=3, max_row=50, values_only=True):
        no, text = row[0], row[1]
        if no is None or text is None:
            continue
        try:
            num = int(no)
        except (TypeError, ValueError):
            continue
        element = " ".join(str(text).split())
        items.append(
            {
                "id": f"item_{num:02d}",
                "label": str(num),
                "name": f"評価項目{num}",
                "element": element,
                "criteria": CRITERIA,
            }
        )
    if not items:
        raise ValueError("評価項目が見つかりません。Excel の形式を確認してください。")
    return items


def build_templates(items: list[dict]) -> tuple[dict, dict]:
    base = {
        "version": "2.0.0",
        "header_fields": [
            {"id": "assignment", "label": "配属"},
            {"id": "job_type", "label": "職種"},
            {"id": "name", "label": "氏名", "readonly": True},
            {"id": "job_title", "label": "役職"},
        ],
        "items": items,
        "text_fields": [],
        "scoring": {"min": 1, "max": 3, "max_total": len(items) * 3},
    }
    self_tpl = {
        **base,
        "id": "self_eval_r8_summer_facility_director",
        "title": "令和8年度 夏季 施設長評価シート（自己評価）",
        "type": "self_evaluation",
        "instructions": (
            "各評価項目について、該当する段階（1～3）を選択してください。"
            "　1：出来ていない　　2：だいたい出来ている（５割以上）　　3：ほぼ出来ている（８割以上）"
        ),
    }
    assess_tpl = {
        **base,
        "id": "assessment_r8_summer_facility_director",
        "title": "令和8年度 夏季 施設長評価シート（考課）",
        "type": "assessment",
        "instructions": (
            "施設長の自己評価を参照し、各項目について該当する段階（1～3）を選択してください。"
            "　二次評価は本部が行います。"
        ),
    }
    return self_tpl, assess_tpl


def write_templates(self_tpl: dict, assess_tpl: dict) -> None:
    for tpl, filename in [
        (self_tpl, "self_evaluation_r8_summer_facility_director.json"),
        (assess_tpl, "assessment_r8_summer_facility_director.json"),
    ]:
        errors = validate_template(tpl)
        if errors:
            raise ValueError(f"{filename}: {errors}")
        path = TEMPLATES_DIR / filename
        path.write_text(json.dumps(tpl, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(f"  wrote {path.name} ({len(tpl['items'])} items)")


def sync_active_period(db, self_tpl: dict, assess_tpl: dict) -> int:
    period = db.query(EvaluationPeriod).filter(EvaluationPeriod.status == PeriodStatus.ACTIVE).first()
    if not period:
        return 0
    updated = 0
    for tpl_id, content in [
        (period.facility_director_self_eval_template_id, self_tpl),
        (period.facility_director_assessment_template_id, assess_tpl),
    ]:
        if not tpl_id:
            continue
        row = db.get(FormTemplate, tpl_id)
        if row:
            row.content = content
            row.version = content["version"]
            row.name = content["title"]
            updated += 1
    db.commit()
    return updated


def main():
    excel_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_EXCEL
    if not excel_path.exists():
        print(f"Excel が見つかりません: {excel_path}", file=sys.stderr)
        sys.exit(1)

    items = parse_items(excel_path)
    self_tpl, assess_tpl = build_templates(items)
    print(f"施設長評価シート: {excel_path.name}")
    write_templates(self_tpl, assess_tpl)

    migrate_schema()
    db = SessionLocal()
    n = sync_active_period(db, self_tpl, assess_tpl)
    if n:
        print(f"  active period DB templates updated: {n}")


if __name__ == "__main__":
    main()
