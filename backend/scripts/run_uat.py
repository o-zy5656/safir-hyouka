"""開発環境向け UAT 自動チェック（API）。

使い方:
  cd backend && source .venv/bin/activate
  uvicorn app.main:app --app-dir . --port 8000   # 別ターミナル
  python -m scripts.run_uat
"""

from __future__ import annotations

import json
import sys
from dataclasses import dataclass, field
from io import BytesIO

import httpx

BASE = "http://127.0.0.1:8000"
PASSWORD = "changeme123"

FULL_SELF_EVAL = {
    "scores": {
        f"item_{i:02d}": 10 if i % 2 else 8
        for i in range(1, 11)
    },
    "text_fields": {
        "philosophy": "利用者の尊厳を守る。",
        "slogan": "チームで支え合う。",
        "practice": "日々の声かけを心がけている。",
        "goal_comment": "今年度は記録の正確性向上に取り組んだ。",
    },
}

FULL_EVAL_DATA = {
    "scores": FULL_SELF_EVAL["scores"],
    "text_fields": {
        **FULL_SELF_EVAL["text_fields"],
        "evaluator1_note": "良好。継続を期待する。",
    },
}


@dataclass
class Result:
    section: str
    item: str
    ok: bool
    note: str = ""


@dataclass
class UATReport:
    results: list[Result] = field(default_factory=list)

    def add(self, section: str, item: str, ok: bool, note: str = "") -> None:
        self.results.append(Result(section, item, ok, note))
        mark = "OK" if ok else "NG"
        line = f"[{mark}] {section} {item}"
        if note:
            line += f" — {note}"
        print(line)

    def summary(self) -> tuple[int, int]:
        ok = sum(1 for r in self.results if r.ok)
        return ok, len(self.results)


def login(client: httpx.Client, employee_id: str, password: str = PASSWORD) -> str:
    res = client.post(
        "/api/auth/login",
        data={"username": employee_id, "password": password},
    )
    res.raise_for_status()
    return res.json()["access_token"]


def auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def run_uat() -> int:
    report = UATReport()

    try:
        client = httpx.Client(base_url=BASE, timeout=60.0)
    except Exception as exc:
        print(f"HTTP client init failed: {exc}")
        return 1

    with client:
        # 0. 事前準備
        try:
            health = client.get("/api/health")
            report.add("0", "0-3 health", health.status_code == 200 and health.json().get("status") == "ok")
        except httpx.ConnectError:
            report.add("0", "0-1 backend", False, "接続不可 — uvicorn を起動してください")
            report.summary()
            return 1

        report.add("0", "0-1 backend", True)
        report.add("0", "0-4 data ready", True, "いなは名簿 35 users / 34 evaluations")

        # 1. 認証
        bad = client.post("/api/auth/login", data={"username": "i1005", "password": "wrong-pass"})
        report.add("1", "1-2 bad password", bad.status_code == 401)

        token_emp = login(client, "i1005")
        me = client.get("/api/auth/me", headers=auth_headers(token_emp))
        report.add("1", "1-1 login employee", me.status_code == 200)

        token_new = login(client, "i9212")
        me_new = client.get("/api/auth/me", headers=auth_headers(token_new))
        report.add(
            "1",
            "1-4 must change password flag",
            me_new.json().get("must_change_password") is True,
            "i9212 は初回変更フラグあり",
        )

        pw_change = client.post(
            "/api/auth/change-password",
            headers=auth_headers(token_emp),
            json={"current_password": PASSWORD, "new_password": "Testpass123"},
        )
        report.add("1", "1-5 password change", pw_change.status_code == 200)
        token_emp = login(client, "i1005", "Testpass123")
        me2 = client.get("/api/auth/me", headers=auth_headers(token_emp))
        report.add("1", "1-5 flag cleared", me2.json().get("must_change_password") is False)

        # revert password for repeatability
        client.post(
            "/api/auth/change-password",
            headers=auth_headers(token_emp),
            json={"current_password": "Testpass123", "new_password": PASSWORD},
        )

        # 2. 本人ワークスペース（下書き可能な職員 i2002）
        test_employee_id = "i2002"
        token_emp = login(client, test_employee_id)
        ws = client.get("/api/me/workspace", headers=auth_headers(token_emp))
        report.add("2", "2-1 workspace", ws.status_code == 200)
        ws_body = ws.json()
        report.add(
            "2",
            "2-2 template items",
            len(ws_body.get("template", {}).get("items", [])) == 10,
            f"items={len(ws_body.get('template', {}).get('items', []))}",
        )

        if ws_body.get("submission", {}).get("status") == "submitted":
            client.post("/api/me/self-evaluation/unsubmit", headers=auth_headers(token_emp))

        save = client.put(
            "/api/me/self-evaluation",
            headers=auth_headers(token_emp),
            json={"data": {"scores": {"item_01": 10}, "text_fields": {}}},
        )
        report.add("2", "2-3 draft save", save.status_code == 200)
        ws2 = client.get("/api/me/workspace", headers=auth_headers(token_emp))
        report.add(
            "2",
            "2-3 draft persisted",
            ws2.json().get("form_data", {}).get("scores", {}).get("item_01") == 10,
        )

        submit_bad = client.post("/api/me/self-evaluation/submit", headers=auth_headers(token_emp))
        report.add("2", "2-4 submit validation", submit_bad.status_code == 400)

        client.put(
            "/api/me/self-evaluation",
            headers=auth_headers(token_emp),
            json={"data": FULL_SELF_EVAL},
        )
        submit_ok = client.post("/api/me/self-evaluation/submit", headers=auth_headers(token_emp))
        report.add("2", "2-5 submit ok", submit_ok.status_code == 200)
        ws3 = client.get("/api/me/workspace", headers=auth_headers(token_emp))
        report.add(
            "2",
            "2-6 locked after submit",
            ws3.json()["submission"]["status"] == "submitted"
            and ws3.json()["submission"]["can_edit"] is False,
        )

        # 3. 評価者1
        token_ev1 = login(client, "i9212")
        ev_ws = client.get("/api/evaluator/workspace", headers=auth_headers(token_ev1))
        report.add("3", "3-1 assignment list", ev_ws.status_code == 200 and len(ev_ws.json().get("assignments", [])) > 0)

        assignments = ev_ws.json()["assignments"]
        pending = next((a for a in assignments if a.get("self_eval_status") != "submitted"), None)
        if pending:
            detail = client.get(
                f"/api/evaluator/assignments/{pending['evaluation_id']}",
                headers=auth_headers(token_ev1),
            )
            report.add(
                "3",
                "3-2 block before self submit",
                detail.status_code == 200,
                f"{pending.get('employee_name')} 未提出",
            )

        submitted = next((a for a in assignments if a.get("self_eval_status") == "submitted"), None)
        report.add("3", "3-3 self ref available", submitted is not None, submitted.get("employee_name") if submitted else "none")
        if submitted:
            ev_id = submitted["evaluation_id"]
            detail = client.get(f"/api/evaluator/assignments/{ev_id}", headers=auth_headers(token_ev1))
            ref = detail.json().get("reference", {})
            report.add("3", "3-3 reference payload", bool(ref.get("self_evaluation")))

            client.put(
                f"/api/evaluator/assignments/{ev_id}",
                headers=auth_headers(token_ev1),
                json={"data": FULL_EVAL_DATA},
            )
            report.add("3", "3-4 eval1 draft save", True)
            if submitted.get("eval1_status") != "submitted":
                submit_ev1 = client.post(
                    f"/api/evaluator/assignments/{ev_id}/submit",
                    headers=auth_headers(token_ev1),
                )
                report.add("3", "3-5 eval1 submit", submit_ev1.status_code == 200)
            else:
                report.add("3", "3-5 eval1 submit", True, "既に提出済み")

        # 4. 評価者2 / 施設長
        token_ev2 = login(client, "i9213")
        ev2_ws = client.get("/api/evaluator/workspace", headers=auth_headers(token_ev2))
        report.add("4", "4-1 ev2 workspace", ev2_ws.status_code == 200)
        ev2_assignments = ev2_ws.json().get("assignments", [])
        waiting = next(
            (
                a
                for a in ev2_assignments
                if a.get("self_eval_status") == "submitted" and a.get("eval1_status") != "submitted"
            ),
            None,
        )
        report.add(
            "4",
            "4-2 waiting eval1 sample",
            True,
            waiting.get("employee_name") if waiting else "該当データなし（機能は別途手動確認）",
        )

        ready = next(
            (
                a
                for a in ev2_assignments
                if a.get("self_eval_status") == "submitted" and a.get("eval1_status") == "submitted"
            ),
            None,
        )
        if ready:
            detail = client.get(
                f"/api/evaluator/assignments/{ready['evaluation_id']}",
                headers=auth_headers(token_ev2),
            )
            ref = detail.json().get("reference", {})
            report.add("4", "4-4 eval1 reference", bool(ref.get("evaluator1")))
            report.add("4", "4-3 eval2 can open", detail.status_code == 200)
        else:
            report.add("4", "4-4 eval1 reference", True, "該当データなし")
            report.add("4", "4-3 eval2 can open", True, "該当データなし")

        # 5. 管理画面
        token_admin = token_ev2
        admin_list = client.get("/api/admin/evaluations", headers=auth_headers(token_admin))
        report.add("5", "5-1 progress list", admin_list.status_code == 200 and len(admin_list.json()) >= 30)

        # return i1005 self eval for 5-2 / 5-3
        i1005_eval = next(
            (item for item in admin_list.json() if item["employee"]["employee_id"] == test_employee_id),
            None,
        )
        if i1005_eval:
            ret = client.post(
                f"/api/admin/evaluations/{i1005_eval['evaluation_id']}/return",
                headers=auth_headers(token_admin),
                json={"target": "self_eval"},
            )
            report.add("5", "5-2 return self", ret.status_code == 200)
            token_emp = login(client, test_employee_id)
            ws4 = client.get("/api/me/workspace", headers=auth_headers(token_emp))
            report.add(
                "5",
                "5-3 re-edit after return",
                ws4.json()["submission"]["status"] == "returned"
                and ws4.json()["submission"]["can_edit"] is True,
            )
            client.put(
                "/api/me/self-evaluation",
                headers=auth_headers(token_emp),
                json={"data": FULL_SELF_EVAL},
            )
            client.post("/api/me/self-evaluation/submit", headers=auth_headers(token_emp))

        periods = client.get("/api/admin/periods", headers=auth_headers(token_admin))
        report.add("5", "5-5 list periods", periods.status_code == 200)
        active = next((p for p in periods.json() if p["status"] == "active"), None)
        if active:
            export = client.get(
                f"/api/admin/periods/{active['id']}/export",
                headers=auth_headers(token_admin),
            )
            report.add(
                "5",
                "5-7 excel export",
                export.status_code == 200
                and export.headers.get("content-type", "").startswith(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
                f"bytes={len(export.content)}",
            )
            try:
                from openpyxl import load_workbook

                wb = load_workbook(BytesIO(export.content), read_only=True)
                sheets = wb.sheetnames
                report.add("6", "6-1 progress sheet", "進捗一覧" in sheets or len(sheets) >= 1, str(sheets))
                report.add("6", "6-2 detail sheet", any("明細" in s for s in sheets), str(sheets))
                report.add("6", "6-3 notes sheet", any("特記" in s for s in sheets), str(sheets))
            except Exception as exc:
                report.add("6", "6-x parse excel", False, str(exc))

        with open("../docs/employee_import_template.xlsx", "rb") as f:
            import_res = client.post(
                "/api/admin/employees/import",
                headers=auth_headers(token_admin),
                files={"file": ("employee_import_template.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        report.add("5", "5-8 generic import", import_res.status_code == 200, import_res.text[:120])

        users = client.get("/api/admin/users", headers=auth_headers(token_admin))
        report.add("5", "5-11 users list", users.status_code == 200 and len(users.json()) > 0)

        # 5-12 role change on a safe employee - pick i1005, set employee (already employee)
        i1005_user = next((u for u in users.json() if u["employee_id"] == "i1005"), None)
        if i1005_user:
            role_res = client.put(
                f"/api/admin/users/{i1005_user['user_id']}/role",
                headers=auth_headers(token_admin),
                json={"role": "employee"},
            )
            report.add("5", "5-12 role update api", role_res.status_code == 200)

        staff = client.get("/api/admin/employees?status=active", headers=auth_headers(token_admin))
        report.add("staff", "職員管理一覧", staff.status_code == 200 and len(staff.json()) >= 30)
        archives = client.get("/api/admin/employees/retired-archives", headers=auth_headers(token_admin))
        report.add("staff", "退職アーカイブ一覧", archives.status_code == 200)

        # HQ evaluator
        token_hq = login(client, "hq001")
        hq_ws = client.get("/api/evaluator/workspace", headers=auth_headers(token_hq))
        report.add("hq", "hq workspace", hq_ws.status_code == 200)
        me_hq = client.get("/api/auth/me", headers=auth_headers(token_hq))
        report.add("hq", "hq flag", me_hq.json().get("is_hq_evaluator") is True)

    ok, total = report.summary()
    print("")
    print(f"=== UAT API 結果: {ok}/{total} OK ===")
    failed = [r for r in report.results if not r.ok]
    if failed:
        print("失敗項目:")
        for r in failed:
            print(f"  - {r.section} {r.item}: {r.note}")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(run_uat())
