"""名簿 Excel（Dropbox 形式）の取込。"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import AuditLog, Employee, EmploymentStatus, Evaluation, EvaluationPeriod, User, UserRole
from app.services.bonus_amounts import default_fiscal_year, ensure_bonus_store_for_facility
from app.services.facilities import (
    get_enabled_facility,
    resolve_facility_for_assignment,
    validate_evaluator_facility,
)
from app.services.form_profiles import FACILITY_DIRECTOR_TITLE
from app.services.hq_evaluator import ensure_hq_evaluator

ROSTER_HEADER_ALIASES = {
    "社員ID": "employee_id",
    "氏名": "name",
    "所属施設": "facility",
    "所属部門": "department",
    "職種": "job_type",
    "役職": "job_title",
    "在籍状態": "employment_status",
    "入社日": "hire_date",
    "一次評価者社員ID": "evaluator1_employee_id",
    "二次評価者社員ID": "evaluator2_employee_id",
    "本部確認対象": "hq_review",
}


@dataclass
class RosterImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    users_created: int = 0
    evaluations_created: int = 0
    roles_updated: int = 0
    errors: list[str] = field(default_factory=list)


def _cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_hire_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip().replace("/", "-").replace(".", "-")
    parts = [p for p in text.split("-") if p]
    try:
        if len(parts) == 3:
            y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
            return date(y, m, d)
    except ValueError:
        return None
    return None


def _years_of_service(hire: Optional[date]) -> int:
    if not hire:
        return 0
    today = date.today()
    years = today.year - hire.year
    if (today.month, today.day) < (hire.month, hire.day):
        years -= 1
    return max(years, 0)


def _normalize_name(name: str) -> str:
    return name.replace("\u3000", " ").strip()


def _find_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    for idx, row in enumerate(rows):
        mapping: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            label = _cell_value(cell)
            if label in ROSTER_HEADER_ALIASES:
                mapping[ROSTER_HEADER_ALIASES[label]] = col_idx
        if "employee_id" in mapping and "name" in mapping:
            return idx, mapping
    raise ValueError("名簿のヘッダー行が見つかりません（社員ID・氏名が必要です）")


def _parse_roster_rows(
    content: bytes,
    facility_filter: Optional[str],
    active_only: bool,
) -> tuple[list[dict[str, str]], list[str]]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_idx, col_map = _find_header_row(rows)

    parsed: list[dict[str, str]] = []
    errors: list[str] = []
    for row_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not row:
            continue

        def get(key: str) -> str:
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return ""
            return _cell_value(row[idx])

        employee_id = get("employee_id")
        if not employee_id:
            continue

        facility = get("facility")
        status = get("employment_status")
        if facility_filter and facility_filter not in facility:
            continue
        if active_only and status and status != "在職":
            continue

        name = _normalize_name(get("name"))
        if not name:
            errors.append(f"{row_idx}行目 ({employee_id}): 氏名が空です")
            continue

        hire = _parse_hire_date(row[col_map["hire_date"]] if "hire_date" in col_map else None)
        parsed.append(
            {
                "employee_id": employee_id,
                "name": name,
                "facility": facility or facility_filter or "",
                "department": get("department"),
                "job_type": get("job_type") or "—",
                "job_title": get("job_title") or "一般",
                "employment_status": status,
                "hire_date": hire.isoformat() if hire else "",
                "years_of_service": str(_years_of_service(hire)),
                "evaluator1_employee_id": get("evaluator1_employee_id"),
                "evaluator2_employee_id": get("evaluator2_employee_id"),
                "hq_review": get("hq_review"),
            }
        )
    return parsed, errors


def _resolve_role(
    employee_id: str,
    job_title: str,
    eval1_refs: set[str],
    eval2_refs: set[str],
    leader_titles: set[str],
) -> UserRole:
    """評価者2=施設長、評価者1=リーダークラス（＋名簿上の一次評価者）。"""
    if job_title == "施設長" or employee_id in eval2_refs:
        return UserRole.EVALUATOR2
    if job_title in leader_titles or employee_id in eval1_refs:
        return UserRole.EVALUATOR1
    return UserRole.EMPLOYEE


def _resolve_is_admin(
    employee_id: str,
    job_title: str,
    admin_titles: set[str],
    admin_ids: set[str],
) -> bool:
    """施設長は評価者2と同時に管理画面へもアクセス可能。"""
    if employee_id in admin_ids:
        return True
    if job_title in admin_titles:
        return True
    return job_title == "施設長"


def resolve_roster_facility_filter(
    *,
    facility_key: Optional[str] = None,
    facility_label: Optional[str] = None,
) -> Optional[str]:
    """名簿取込の施設絞り込み。None は全施設。"""
    if facility_key:
        key = facility_key.strip().lower()
        if key in {"", "all", "*"}:
            return None
        return get_enabled_facility(facility_key).assignment_match
    if facility_label:
        label = facility_label.strip()
        if label in {"", "全施設", "all", "*"}:
            return None
        return label
    return None


def import_roster_from_excel(
    db: Session,
    content: bytes,
    admin_user: Optional[User],
    password_hash_fn,
    default_password: str,
    active_period: Optional[EvaluationPeriod],
    facility_filter: Optional[str] = None,
    facility_key: Optional[str] = None,
    active_only: bool = True,
    admin_job_titles: Optional[set[str]] = None,
    admin_employee_ids: Optional[set[str]] = None,
    leader_titles: Optional[set[str]] = None,
) -> RosterImportResult:
    resolved_filter = resolve_roster_facility_filter(
        facility_key=facility_key,
        facility_label=facility_filter,
    )
    rows, parse_errors = _parse_roster_rows(content, resolved_filter, active_only)
    result = RosterImportResult(errors=list(parse_errors))
    if not rows and parse_errors:
        return result
    if not rows:
        scope = resolved_filter or "全施設"
        result.errors.append(f"対象データがありません（施設: {scope}）")
        return result

    admin_titles = admin_job_titles or {"施設長"}
    admin_ids = {x.strip() for x in (admin_employee_ids or set()) if x.strip()}
    leaders = leader_titles or {"リーダー", "サブリーダー"}

    eval1_refs = {r["evaluator1_employee_id"] for r in rows if r.get("evaluator1_employee_id")}
    eval2_refs = {r["evaluator2_employee_id"] for r in rows if r.get("evaluator2_employee_id")}

    employee_cache: dict[str, Employee] = {
        emp.employee_id: emp for emp in db.query(Employee).all()
    }

    hq_employee = ensure_hq_evaluator(db, password_hash_fn, default_password)
    employee_cache[hq_employee.employee_id] = hq_employee

    def get_or_create_user(employee_id: str) -> tuple[User, bool]:
        user = db.query(User).filter(User.employee_id == employee_id).first()
        if user:
            return user, False
        user = User(
            employee_id=employee_id,
            password_hash=password_hash_fn(default_password),
            role=UserRole.EMPLOYEE,
            must_change_password=True,
        )
        db.add(user)
        db.flush()
        return user, True

    for record in rows:
        employee_id = record["employee_id"]
        years = int(record["years_of_service"])
        assignment = record["facility"]
        if record.get("department"):
            assignment = f"{record['facility']} {record['department']}".strip()

        existing = employee_cache.get(employee_id)
        if existing:
            existing.name = record["name"]
            existing.assignment = assignment
            existing.job_type = record["job_type"]
            existing.job_title = record["job_title"]
            existing.years_of_service = years
            existing.employment_status = EmploymentStatus.ACTIVE
            existing.retired_at = None
            result.updated += 1
        else:
            user, created_user = get_or_create_user(employee_id)
            if created_user:
                result.users_created += 1
            existing = Employee(
                employee_id=employee_id,
                name=record["name"],
                assignment=assignment,
                job_type=record["job_type"],
                job_title=record["job_title"],
                years_of_service=years,
                user_id=user.id,
                employment_status=EmploymentStatus.ACTIVE,
            )
            db.add(existing)
            db.flush()
            employee_cache[employee_id] = existing
            result.created += 1

        if not existing.user_id:
            user, created_user = get_or_create_user(employee_id)
            if created_user:
                result.users_created += 1
            existing.user_id = user.id

    for record in rows:
        employee = employee_cache.get(record["employee_id"])
        if not employee:
            continue
        ev1_id = record.get("evaluator1_employee_id", "")
        ev2_id = record.get("evaluator2_employee_id", "")
        job_title = record.get("job_title", "")

        if ev1_id:
            ev1 = employee_cache.get(ev1_id)
            if not ev1:
                result.errors.append(f"{record['employee_id']}: 一次評価者 ({ev1_id}) が名簿にいません")
            else:
                warning = validate_evaluator_facility(employee, ev1, role_label="第1評価者")
                if warning:
                    result.errors.append(warning)
                employee.evaluator1_id = ev1.id
        else:
            employee.evaluator1_id = None

        if job_title == FACILITY_DIRECTOR_TITLE:
            employee.evaluator1_id = None
            employee.evaluator2_id = hq_employee.id
        elif ev2_id:
            ev2 = employee_cache.get(ev2_id)
            if not ev2:
                result.errors.append(f"{record['employee_id']}: 二次評価者 ({ev2_id}) が名簿にいません")
            else:
                warning = validate_evaluator_facility(employee, ev2, role_label="第2評価者")
                if warning:
                    result.errors.append(warning)
                employee.evaluator2_id = ev2.id
        else:
            employee.evaluator2_id = None

    for record in rows:
        user = db.query(User).filter(User.employee_id == record["employee_id"]).first()
        if not user:
            continue
        new_role = _resolve_role(
            record["employee_id"],
            record["job_title"],
            eval1_refs,
            eval2_refs,
            leaders,
        )
        is_admin = _resolve_is_admin(
            record["employee_id"],
            record["job_title"],
            admin_titles,
            admin_ids,
        )
        if new_role != user.role:
            user.role = new_role
            result.roles_updated += 1
        if is_admin != user.is_admin:
            user.is_admin = is_admin
            result.roles_updated += 1

    if active_period:
        existing_eval_employee_ids = {
            employee_id
            for (employee_id,) in db.query(Evaluation.employee_id)
            .filter(Evaluation.period_id == active_period.id)
            .all()
        }
        for record in rows:
            employee = employee_cache.get(record["employee_id"])
            if not employee or employee.id in existing_eval_employee_ids:
                continue
            db.add(Evaluation(period_id=active_period.id, employee_id=employee.id))
            result.evaluations_created += 1

    if admin_user:
        db.add(
            AuditLog(
                user_id=admin_user.id,
                action="import_roster",
                target_type="employees",
                target_id=active_period.id if active_period else None,
                detail={
                    "facility": facility_filter,
                    "created": result.created,
                    "updated": result.updated,
                    "roles_updated": result.roles_updated,
                },
            )
        )

    affected_facility_keys: set[str] = set()
    for record in rows:
        assignment = record["facility"]
        if record.get("department"):
            assignment = f"{record['facility']} {record['department']}".strip()
        facility = resolve_facility_for_assignment(assignment)
        if facility:
            affected_facility_keys.add(facility.key)

    fiscal_year = active_period.fiscal_year if active_period else default_fiscal_year()
    for facility_key in affected_facility_keys:
        ensure_bonus_store_for_facility(facility_key, fiscal_year)

    db.commit()
    return result
