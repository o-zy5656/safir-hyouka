"""考課期間の Excel 出力（進捗一覧 + 項目明細）。"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.models import Employee, Evaluation, EvaluationPeriod, SubmissionStatus

STATUS_LABELS = {
    SubmissionStatus.PENDING: "未着手",
    SubmissionStatus.DRAFT: "下書き",
    SubmissionStatus.SUBMITTED: "提出済",
    SubmissionStatus.RETURNED: "差戻",
}


def _status_label(status: SubmissionStatus) -> str:
    return STATUS_LABELS.get(status, status.value)


def _sum_scores(data: dict[str, Any]) -> Optional[int]:
    scores = (data or {}).get("scores") or {}
    if not scores:
        return None
    total = 0
    for value in scores.values():
        if value in (None, ""):
            return None
        total += int(value)
    return total


def build_period_export(
    db: Session,
    period: EvaluationPeriod,
    self_template: dict,
    assessment_template: dict,
) -> tuple[bytes, str]:
    evaluations = db.query(Evaluation).filter(Evaluation.period_id == period.id).all()
    employee_map = {emp.id: emp for emp in db.query(Employee).all()}

    wb = Workbook()
    summary = wb.active
    summary.title = "進捗一覧"
    summary.append(
        [
            "社員ID",
            "氏名",
            "配属",
            "職種",
            "勤続年数",
            "自己評価",
            "評価者1",
            "評価者2",
            "自己評価合計",
            "評1合計",
            "評2合計",
        ]
    )
    for cell in summary[1]:
        cell.font = Font(bold=True)

    detail = wb.create_sheet("項目明細")
    detail.append(
        [
            "社員ID",
            "氏名",
            "項目",
            "項目名",
            "自己採点",
            "評1採点",
            "評2採点",
        ]
    )
    for cell in detail[1]:
        cell.font = Font(bold=True)

    items = self_template.get("items", [])
    for evaluation in evaluations:
        employee = employee_map.get(evaluation.employee_id)
        if not employee:
            continue

        self_total = _sum_scores(evaluation.self_eval_data)
        eval1_total = _sum_scores(evaluation.eval1_data)
        eval2_total = _sum_scores(evaluation.eval2_data)

        summary.append(
            [
                employee.employee_id,
                employee.name,
                employee.assignment,
                employee.job_type,
                employee.years_of_service,
                _status_label(evaluation.self_eval_status),
                _status_label(evaluation.eval1_status),
                _status_label(evaluation.eval2_status),
                self_total,
                eval1_total,
                eval2_total,
            ]
        )

        self_scores = (evaluation.self_eval_data or {}).get("scores") or {}
        eval1_scores = (evaluation.eval1_data or {}).get("scores") or {}
        eval2_scores = (evaluation.eval2_data or {}).get("scores") or {}
        for item in items:
            item_id = item["id"]
            detail.append(
                [
                    employee.employee_id,
                    employee.name,
                    item.get("label", item_id),
                    item.get("name", ""),
                    self_scores.get(item_id),
                    eval1_scores.get(item_id),
                    eval2_scores.get(item_id),
                ]
            )

    notes = wb.create_sheet("特記事項")
    notes.append(["社員ID", "氏名", "項目", "内容"])
    for cell in notes[1]:
        cell.font = Font(bold=True)

    self_text_fields = self_template.get("text_fields") or []
    assess_text_fields = assessment_template.get("text_fields") or []
    for evaluation in evaluations:
        employee = employee_map.get(evaluation.employee_id)
        if not employee:
            continue
        self_text = (evaluation.self_eval_data or {}).get("text_fields") or {}
        eval1_text = (evaluation.eval1_data or {}).get("text_fields") or {}
        eval2_text = (evaluation.eval2_data or {}).get("text_fields") or {}

        for field in self_text_fields:
            value = self_text.get(field["id"], "")
            if value:
                notes.append([employee.employee_id, employee.name, field.get("label", field["id"]), value])

        for field in assess_text_fields:
            if field.get("readonly"):
                continue
            field_id = field["id"]
            role = field.get("role")
            if role == "evaluator1":
                value = eval1_text.get(field_id, "")
            elif role == "evaluator2":
                value = eval2_text.get(field_id, "")
            else:
                value = eval1_text.get(field_id) or eval2_text.get(field_id) or ""
            if value:
                notes.append([employee.employee_id, employee.name, field.get("label", field_id), value])

    buffer = BytesIO()
    wb.save(buffer)
    safe_name = period.name.replace("/", "_").replace(" ", "_")
    filename = f"hyouka_{safe_name}.xlsx"
    return buffer.getvalue(), filename
