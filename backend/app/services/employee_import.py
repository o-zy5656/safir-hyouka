"""社員Excel一括取込（PLAN.md 8章の列形式 + 任意のロール列）。"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import AuditLog, Employee, EmploymentStatus, Evaluation, EvaluationPeriod, User, UserRole


IMPORT_COLUMNS = {
    "社員ID": "employee_id",
    "氏名": "name",
    "配属": "assignment",
    "職種": "job_type",
    "勤続年数": "years_of_service",
    "評価者1社員ID": "evaluator1_employee_id",
    "評価者2社員ID": "evaluator2_employee_id",
    "ロール": "role",
}

REQUIRED_COLUMNS = {"社員ID", "氏名", "配属", "職種", "勤続年数", "評価者1社員ID"}

ROLE_ALIASES: dict[str, UserRole] = {
    "employee": UserRole.EMPLOYEE,
    "本人": UserRole.EMPLOYEE,
    "被評価者": UserRole.EMPLOYEE,
    "evaluator1": UserRole.EVALUATOR1,
    "評価者1": UserRole.EVALUATOR1,
    "評1": UserRole.EVALUATOR1,
    "evaluator2": UserRole.EVALUATOR2,
    "評価者2": UserRole.EVALUATOR2,
    "評2": UserRole.EVALUATOR2,
    "admin": UserRole.ADMIN,
    "管理者": UserRole.ADMIN,
}


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    users_created: int = 0
    evaluations_created: int = 0
    roles_updated: int = 0
    errors: list[str] = field(default_factory=list)


def parse_role_value(value: str) -> Optional[UserRole]:
    normalized = value.strip().lower()
    if not normalized:
        return None
    if normalized in ROLE_ALIASES:
        return ROLE_ALIASES[normalized]
    for key, role in ROLE_ALIASES.items():
        if key.lower() == normalized:
            return role
    return None


def _cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_rows(content: bytes) -> tuple[list[dict[str, str]], list[str], bool]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        return [], ["Excelにデータがありません"], False

    headers = [_cell_value(cell) for cell in header_row]
    has_role_column = "ロール" in headers
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        return [], [f"必須列がありません: {', '.join(sorted(missing))}"], has_role_column

    parsed: list[dict[str, str]] = []
    errors: list[str] = []
    for row_idx, row in enumerate(rows, start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record: dict[str, str] = {}
        for col_idx, header in enumerate(headers):
            if header not in IMPORT_COLUMNS:
                continue
            key = IMPORT_COLUMNS[header]
            record[key] = _cell_value(row[col_idx] if col_idx < len(row) else "")

        if not record.get("employee_id"):
            errors.append(f"{row_idx}行目: 社員IDが空です")
            continue
        if not record.get("name"):
            errors.append(f"{row_idx}行目 ({record['employee_id']}): 氏名が空です")
            continue
        try:
            record["years_of_service"] = str(int(record.get("years_of_service") or "0"))
        except ValueError:
            errors.append(f"{row_idx}行目 ({record['employee_id']}): 勤続年数が不正です")
            continue
        if record.get("role"):
            if parse_role_value(record["role"]) is None:
                errors.append(
                    f"{row_idx}行目 ({record['employee_id']}): ロール「{record['role']}」が不正です"
                )
                continue
        parsed.append(record)
    return parsed, errors, has_role_column


def _get_or_create_user(
    db: Session,
    employee_id: str,
    password_hash_fn,
    default_password: str,
) -> tuple[User, bool]:
    user = db.query(User).filter(User.employee_id == employee_id).first()
    if user:
        return user, False
    user = User(
        employee_id=employee_id,
        password_hash=password_hash_fn(default_password),
        role=UserRole.EMPLOYEE,
        must_change_password=True,
    )
    db.add(user)
    db.flush()
    return user, True


def _apply_user_roles(
    db: Session,
    rows: list[dict[str, str]],
    has_role_column: bool,
    result: ImportResult,
) -> None:
    eval1_refs = {record["evaluator1_employee_id"] for record in rows if record.get("evaluator1_employee_id")}
    eval2_refs = {record["evaluator2_employee_id"] for record in rows if record.get("evaluator2_employee_id")}

    explicit_roles: dict[str, UserRole] = {}
    if has_role_column:
        for record in rows:
            role_text = record.get("role", "")
            if role_text:
                parsed = parse_role_value(role_text)
                if parsed:
                    explicit_roles[record["employee_id"]] = parsed

    target_ids = {record["employee_id"] for record in rows} | eval1_refs | eval2_refs
    for employee_id in target_ids:
        user = db.query(User).filter(User.employee_id == employee_id).first()
        if not user or user.role == UserRole.ADMIN:
            continue

        new_role: Optional[UserRole] = None
        if employee_id in explicit_roles:
            new_role = explicit_roles[employee_id]
        elif not has_role_column:
            if employee_id in eval2_refs:
                new_role = UserRole.EVALUATOR2
            elif employee_id in eval1_refs:
                new_role = UserRole.EVALUATOR1

        if new_role and new_role != user.role:
            user.role = new_role
            result.roles_updated += 1


def import_employees_from_excel(
    db: Session,
    content: bytes,
    admin_user: User,
    password_hash_fn,
    default_password: str,
    active_period: Optional[EvaluationPeriod],
) -> ImportResult:
    rows, parse_errors, has_role_column = _parse_rows(content)
    result = ImportResult(errors=list(parse_errors))
    if parse_errors and not rows:
        return result

    employee_cache: dict[str, Employee] = {
        emp.employee_id: emp for emp in db.query(Employee).all()
    }

    for record in rows:
        employee_id = record["employee_id"]
        years = int(record["years_of_service"])
        existing = employee_cache.get(employee_id)
        if existing:
            existing.name = record["name"]
            existing.assignment = record["assignment"]
            existing.job_type = record["job_type"]
            existing.years_of_service = years
            existing.employment_status = EmploymentStatus.ACTIVE
            existing.retired_at = None
            result.updated += 1
        else:
            user, created_user = _get_or_create_user(
                db, employee_id, password_hash_fn, default_password
            )
            if created_user:
                result.users_created += 1
            existing = Employee(
                employee_id=employee_id,
                name=record["name"],
                assignment=record["assignment"],
                job_type=record["job_type"],
                years_of_service=years,
                user_id=user.id,
                employment_status=EmploymentStatus.ACTIVE,
            )
            db.add(existing)
            db.flush()
            employee_cache[employee_id] = existing
            result.created += 1

        if not existing.user_id:
            user, created_user = _get_or_create_user(
                db, employee_id, password_hash_fn, default_password
            )
            if created_user:
                result.users_created += 1
            existing.user_id = user.id

    for record in rows:
        employee = employee_cache[record["employee_id"]]
        ev1_id = record.get("evaluator1_employee_id", "")
        ev2_id = record.get("evaluator2_employee_id", "")

        if ev1_id:
            ev1 = employee_cache.get(ev1_id)
            if not ev1:
                result.errors.append(f"{record['employee_id']}: 評価者1 ({ev1_id}) が見つかりません")
            else:
                employee.evaluator1_id = ev1.id
        else:
            employee.evaluator1_id = None

        if ev2_id:
            ev2 = employee_cache.get(ev2_id)
            if not ev2:
                result.errors.append(f"{record['employee_id']}: 評価者2 ({ev2_id}) が見つかりません")
            else:
                employee.evaluator2_id = ev2.id
        else:
            employee.evaluator2_id = None

    _apply_user_roles(db, rows, has_role_column, result)

    if active_period:
        existing_eval_employee_ids = {
            employee_id
            for (employee_id,) in db.query(Evaluation.employee_id)
            .filter(Evaluation.period_id == active_period.id)
            .all()
        }
        for record in rows:
            employee = employee_cache.get(record["employee_id"])
            if (
                not employee
                or employee.employment_status != EmploymentStatus.ACTIVE
                or employee.id in existing_eval_employee_ids
            ):
                continue
            db.add(Evaluation(period_id=active_period.id, employee_id=employee.id))
            result.evaluations_created += 1

    db.add(
        AuditLog(
            user_id=admin_user.id,
            action="import",
            target_type="employees",
            target_id=active_period.id if active_period else None,
            detail={
                "created": result.created,
                "updated": result.updated,
                "users_created": result.users_created,
                "evaluations_created": result.evaluations_created,
                "roles_updated": result.roles_updated,
                "error_count": len(result.errors),
            },
        )
    )
    db.commit()
    return result
