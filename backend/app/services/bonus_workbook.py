"""賞与・昇給 Excel テンプレートへ考課結果を反映する。"""

from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

import msoffcrypto
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import settings
from app.models import Employee, EmploymentStatus, Evaluation, EvaluationPeriod, PeriodStatus, SubmissionStatus
from app.services.facilities import (
    FACILITY_DIRECTORS_BONUS_LABEL,
    build_bonus_preset,
    get_enabled_facility,
    is_facility_directors_bonus_key,
    resolve_facility_for_assignment,
    user_has_global_facility_access,
)
from app.services.bonus_amounts import default_fiscal_year, save_workbook_row_fields
from app.services.form_profiles import evaluation_skips_eval1, is_facility_director

LOW_SCORE_ITEM_IDS = [f"item_{index:02d}" for index in range(1, 6)]
PROMOTION_ITEM_IDS = [f"item_{index:02d}" for index in range(6, 11)]
ITEM_LABELS = {f"item_{index:02d}": label for index, label in enumerate("①②③④⑤⑥⑦⑧⑨⑩", start=1)}
LOW_SELF_THRESHOLD = 8
LOW_OTHER_THRESHOLD = 6
PROMOTION_THRESHOLD = 8
SALARY_RAISE_CORE_TOTAL = 40
SALARY_RAISE_MARK = "○"
RANK_GRADES = ("A", "B", "C", "D", "E")
ROLE_HOLDER_TITLES = frozenset({"リーダー", "サブリーダー", "施設長", "特養管理者"})


def _get_preset(facility_key: str, wb: Any = None) -> dict[str, Any]:
    sheet_names: list[str] = []
    if wb is not None:
        sheet_names = list(wb.sheetnames)
    else:
        opened = _try_open_workbook(data_only=True)
        if opened is not None:
            sheet_names = list(opened.sheetnames)
    return build_bonus_preset(facility_key, sheet_names=sheet_names)


def _load_bonus_rows_from_roster(
    db: Session,
    facility_key: str,
    preset: dict[str, Any],
) -> tuple[list[dict[str, Any]], str]:
    employees = sorted(_collect_facility_employees(db, preset), key=lambda emp: emp.name)
    evaluations = _active_period_evaluations(db)
    rows: list[dict[str, Any]] = []

    for index, employee in enumerate(employees):
        row_number = preset["first_data_row"] + index
        evaluation = evaluations.get(employee.id)
        self_score = None
        eval1_score = None
        eval2_score = None
        final_score = None
        note = ""
        low_self_count = None
        low_other_count = None
        salary_raise = None

        if evaluation:
            totals = _extract_evaluation_totals(employee, evaluation)
            self_score = totals["self_total"]
            eval1_score = totals["eval1_total"] if totals["use_eval1"] else None
            eval2_score = totals["eval2_total"]
            final_score = totals["final_score"]
            note = totals["note"]

        row_data = _row_dict_from_sources(
            row_number=row_number,
            employee=employee,
            evaluation=evaluation,
            self_score=self_score,
            eval1_score=eval1_score,
            eval2_score=eval2_score,
            final_score=final_score,
            low_self_count=low_self_count,
            low_other_count=low_other_count,
            salary_raise=salary_raise,
            rank_order=None,
            rank_grade=None,
            note=note,
        )
        if is_facility_director(employee):
            row_data["low_self_count"] = None
            row_data["low_other_count"] = None
            row_data["salary_raise"] = None
        rows.append(row_data)

    return rows, preset["label"]


@dataclass
class BonusReflectResult:
    facility: str
    updated_rows: int = 0
    matched_employees: list[str] = field(default_factory=list)
    unmatched_employees: list[str] = field(default_factory=list)
    unmatched_excel_names: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def format_bonus_name(name: str) -> str:
    """賞与 Excel 形式（姓と名の間は全角スペース）。"""
    text = unicodedata.normalize("NFKC", name or "").strip()
    if not text:
        return ""
    if "\u3000" in text:
        return text
    parts = text.split()
    if len(parts) >= 2:
        return "\u3000".join(parts)
    return text


def bonus_job_title_label(employee: Employee) -> str:
    """名簿（自己評価システム）の職種・役職を賞与表の表示名に変換。"""
    job_title = (employee.job_title or "").strip()
    job_type = (employee.job_type or "").strip()

    if job_title == "施設長":
        return "施設長"
    if job_title == "特養管理者":
        return "特養管理者"
    if job_type == "介護支援専門員":
        return "生活相談員"
    if job_title in {"リーダー", "サブリーダー"}:
        return job_title
    if job_type in {"看護師", "准看護師"}:
        return "看護職員"
    if job_type == "管理栄養士":
        return "厨房"
    if job_type == "事務":
        return "事務員"
    if job_type == "ケアアシスト":
        return "ケアアシスト\u3000P"
    if job_title == "一般" or job_title == "—":
        if "P" in job_type.upper() or job_type.endswith("P"):
            return "介護\u3000P"
        return "介護一般"
    return job_title or job_type or "一般"


def distribute_grade_counts(
    total: int,
    percentages: list[tuple[str, float]],
) -> list[tuple[str, int]]:
    """人数割合（評価基準.xlsx）を各ランクの人数に按分（最大剰余法）。"""
    if total <= 0:
        return [(grade, 0) for grade, _ in percentages]

    exact = [ratio * total for _, ratio in percentages]
    counts = [int(value) for value in exact]
    remainder = total - sum(counts)
    if remainder > 0:
        fractions = sorted(
            ((exact[index] - counts[index], index) for index in range(len(percentages))),
            reverse=True,
        )
        for offset in range(remainder):
            counts[fractions[offset][1]] += 1
    return [(percentages[index][0], counts[index]) for index in range(len(percentages))]


def grade_for_rank(
    rank: int,
    total: int,
    percentages: Optional[list[tuple[str, float]]] = None,
) -> str:
    """考課点順位（1始まり）から A〜E を決定。"""
    rules = percentages or settings.bonus_rank_grade_percentages
    if not rules or total <= 0:
        return "E"
    cursor = 0
    for grade, count in distribute_grade_counts(total, rules):
        cursor += count
        if rank <= cursor:
            return grade
    return "E"


def _apply_employee_identity(ws, row: int, preset: dict[str, Any], employee: Employee) -> None:
    ws.cell(row, preset["name_col"]).value = format_bonus_name(employee.name)
    ws.cell(row, preset["job_col"]).value = bonus_job_title_label(employee)


def _apply_auto_rank_orders(
    ws,
    preset: dict[str, Any],
    row_final_scores: list[tuple[int, int]],
) -> None:
    if not row_final_scores:
        return
    ranked = sorted(row_final_scores, key=lambda item: (-item[1], item[0]))
    total = len(ranked)
    for order, (row, _final_score) in enumerate(ranked, start=1):
        ws.cell(row, preset["rank_order_col"]).value = order
        if settings.bonus_auto_rank_grade:
            ws.cell(row, preset["rank_grade_col"]).value = grade_for_rank(order, total)


def normalize_person_name(name: str) -> str:
    text = unicodedata.normalize("NFKC", name or "")
    text = text.replace("渡邉", "渡邊")
    return re.sub(r"\s+", "", text)


def _sum_scores(data: Optional[dict[str, Any]]) -> Optional[int]:
    scores = (data or {}).get("scores") or {}
    if not scores:
        return None
    total = 0
    for value in scores.values():
        if value in (None, ""):
            return None
        total += int(value)
    return total


def _list_low_item_labels(
    scores: Optional[dict[str, Any]],
    item_ids: list[str],
    *,
    threshold: int,
    inclusive_max: bool = False,
) -> list[str]:
    if not scores:
        return []
    labels: list[str] = []
    for item_id in item_ids:
        value = scores.get(item_id)
        if value in (None, ""):
            continue
        score = int(value)
        is_low = score <= threshold if inclusive_max else score < threshold
        if is_low:
            labels.append(ITEM_LABELS[item_id])
    return labels


def _format_item_summary(labels: list[str]) -> Optional[str]:
    if not labels:
        return None
    return f"{''.join(labels)}（{len(labels)}件）"


def _count_low_self_scores(scores: Optional[dict[str, Any]]) -> Optional[int]:
    labels = _list_low_item_labels(
        scores,
        LOW_SCORE_ITEM_IDS,
        threshold=LOW_SELF_THRESHOLD,
    )
    return len(labels) or None


def _count_low_other_scores(scores: Optional[dict[str, Any]]) -> Optional[int]:
    labels = _list_low_item_labels(
        scores,
        LOW_SCORE_ITEM_IDS,
        threshold=LOW_OTHER_THRESHOLD,
        inclusive_max=True,
    )
    return len(labels) or None


def is_role_holder(employee: Optional[Employee]) -> bool:
    return bool(employee and (employee.job_title or "").strip() in ROLE_HOLDER_TITLES)


def _submitted_scores(
    data: Optional[dict[str, Any]],
    status: SubmissionStatus,
) -> dict[str, Any]:
    if status != SubmissionStatus.SUBMITTED:
        return {}
    return (data or {}).get("scores") or {}


def build_evaluation_insights(
    employee: Optional[Employee],
    evaluation: Optional[Evaluation],
) -> dict[str, Any]:
    """考課データからカット対象・昇格参考（読取専用）を算出。"""
    empty: dict[str, Any] = {
        "cut_self_items": None,
        "cut_other_items": None,
        "promotion_reference": None,
        "is_role_holder": is_role_holder(employee),
    }
    if not employee or not evaluation:
        return empty

    if is_facility_director(employee):
        return empty

    self_scores = _submitted_scores(evaluation.self_eval_data, evaluation.self_eval_status)
    skips_eval1 = evaluation_skips_eval1(employee)
    use_eval1 = (
        not skips_eval1
        and evaluation.eval1_status == SubmissionStatus.SUBMITTED
        and _sum_scores(evaluation.eval1_data) is not None
    )
    other_scores = _other_scores_for_rules(evaluation, use_eval1)

    cut_self_labels = _list_low_item_labels(
        self_scores,
        LOW_SCORE_ITEM_IDS,
        threshold=LOW_SELF_THRESHOLD,
    )
    cut_other_labels = _list_low_item_labels(
        other_scores,
        LOW_SCORE_ITEM_IDS,
        threshold=LOW_OTHER_THRESHOLD,
        inclusive_max=True,
    )
    promo_self_labels = _list_low_item_labels(
        self_scores,
        PROMOTION_ITEM_IDS,
        threshold=PROMOTION_THRESHOLD,
    )
    promo_other_labels = _list_low_item_labels(
        other_scores,
        PROMOTION_ITEM_IDS,
        threshold=PROMOTION_THRESHOLD,
    )

    promo_parts: list[str] = []
    if promo_self_labels:
        promo_parts.append(f"自己{''.join(promo_self_labels)}")
    if promo_other_labels:
        promo_parts.append(f"他者{''.join(promo_other_labels)}")

    return {
        "cut_self_items": _format_item_summary(cut_self_labels),
        "cut_other_items": _format_item_summary(cut_other_labels),
        "promotion_reference": " / ".join(promo_parts) if promo_parts else None,
        "is_role_holder": is_role_holder(employee),
    }


def _sum_core_item_scores(scores: Optional[dict[str, Any]]) -> Optional[int]:
    if not scores:
        return None
    total = 0
    for item_id in LOW_SCORE_ITEM_IDS:
        value = scores.get(item_id)
        if value in (None, ""):
            return None
        total += int(value)
    return total


def _salary_raise_mark(other_scores: Optional[dict[str, Any]]) -> Optional[str]:
    """①〜⑤合計 40 点以上なら昇給検討対象。"""
    total = _sum_core_item_scores(other_scores)
    if total is not None and total >= SALARY_RAISE_CORE_TOTAL:
        return SALARY_RAISE_MARK
    return None


def _other_scores_for_rules(evaluation: Evaluation, use_eval1: bool) -> dict[str, Any]:
    eval2_scores = (evaluation.eval2_data or {}).get("scores") or {}
    if eval2_scores:
        return eval2_scores
    if use_eval1:
        return (evaluation.eval1_data or {}).get("scores") or {}
    return {}


def _pick_note(eval1_text: dict, eval2_text: dict) -> str:
    for key in ("evaluator2_note", "evaluator1_note"):
        value = eval2_text.get(key) or eval1_text.get(key) or ""
        if str(value).strip():
            return str(value).strip()
    return ""


def _collect_all_facility_directors(db: Session) -> list[Employee]:
    return (
        db.query(Employee)
        .filter(
            Employee.employment_status == EmploymentStatus.ACTIVE,
            Employee.job_title == "施設長",
        )
        .order_by(Employee.employee_id)
        .all()
    )


def _extract_evaluation_totals(employee: Employee, evaluation: Evaluation) -> dict[str, Any]:
    self_total = None
    eval1_total = None
    eval2_total = None
    self_scores: dict = {}
    other_scores: dict = {}
    note = ""

    if evaluation.self_eval_status == SubmissionStatus.SUBMITTED:
        self_total = _sum_scores(evaluation.self_eval_data)
        self_scores = (evaluation.self_eval_data or {}).get("scores") or {}

    skips_eval1 = evaluation_skips_eval1(employee)
    use_eval1 = (
        not skips_eval1
        and evaluation.eval1_status == SubmissionStatus.SUBMITTED
        and _sum_scores(evaluation.eval1_data) is not None
    )
    if use_eval1:
        eval1_total = _sum_scores(evaluation.eval1_data)
    if evaluation.eval2_status == SubmissionStatus.SUBMITTED:
        eval2_total = _sum_scores(evaluation.eval2_data)

    eval1_text = (evaluation.eval1_data or {}).get("text_fields") or {}
    eval2_text = (evaluation.eval2_data or {}).get("text_fields") or {}
    note = _pick_note(eval1_text, eval2_text)
    other_scores = _other_scores_for_rules(evaluation, use_eval1)
    final_score = _compute_final_score(
        self_total,
        eval1_total if use_eval1 else None,
        eval2_total,
    )
    has_scores = (
        self_total is not None
        or eval1_total is not None
        or eval2_total is not None
        or bool(note)
    )
    return {
        "self_total": self_total,
        "eval1_total": eval1_total,
        "eval2_total": eval2_total,
        "use_eval1": use_eval1,
        "self_scores": self_scores,
        "other_scores": other_scores,
        "note": note,
        "final_score": final_score,
        "has_scores": has_scores,
    }


def _apply_evaluation_for_employee(
    ws,
    row: int,
    preset: dict[str, Any],
    employee: Employee,
    evaluation: Optional[Evaluation],
    result: BonusReflectResult,
    *,
    dry_run: bool,
) -> Optional[int]:
    if not dry_run:
        _apply_employee_identity(ws, row, preset, employee)

    if not evaluation:
        result.warnings.append(f"{employee.name}: 考課データなし（氏名・役職のみ反映）")
        return None

    totals = _extract_evaluation_totals(employee, evaluation)
    if not totals["has_scores"]:
        result.warnings.append(f"{employee.name}: 提出済み考課なし（氏名・役職のみ反映）")
        return None

    if not dry_run:
        _apply_row_values(
            ws,
            row,
            preset,
            self_total=totals["self_total"],
            eval1_total=totals["eval1_total"],
            eval2_total=totals["eval2_total"],
            use_eval1=totals["use_eval1"],
            self_scores=totals["self_scores"],
            other_scores=totals["other_scores"],
            note=totals["note"],
            apply_cut_rules=not is_facility_director(employee),
        )
    return totals["final_score"]


def _row_dict_from_sources(
    *,
    row_number: int,
    employee: Employee,
    evaluation: Optional[Evaluation],
    self_score: Optional[int],
    eval1_score: Optional[int],
    eval2_score: Optional[int],
    final_score: Optional[int],
    low_self_count: Optional[int],
    low_other_count: Optional[int],
    salary_raise: Optional[str],
    rank_order: Optional[int],
    rank_grade: Optional[str],
    note: str,
    bonus_facility_key: Optional[str] = None,
    facility_label: Optional[str] = None,
) -> dict[str, Any]:
    row_data: dict[str, Any] = {
        "row_number": row_number,
        "employee_id": employee.employee_id,
        "name": format_bonus_name(employee.name),
        "job_title": bonus_job_title_label(employee),
        "self_score": self_score,
        "eval1_score": eval1_score,
        "eval2_score": eval2_score,
        "final_score": final_score,
        "low_self_count": low_self_count,
        "low_other_count": low_other_count,
        "salary_raise": salary_raise,
        "rank_order": rank_order,
        "rank_grade": rank_grade,
        "note": note,
        "bonus_facility_key": bonus_facility_key,
        "facility_label": facility_label,
    }
    row_data.update(build_evaluation_insights(employee, evaluation))
    return row_data


def _load_directors_bonus_rows(db: Session) -> tuple[list[dict[str, Any]], str]:
    directors = _collect_all_facility_directors(db)
    evaluations = _active_period_evaluations(db)
    aliases = _load_bonus_aliases()
    wb = _try_open_workbook(data_only=True)
    rows: list[dict[str, Any]] = []

    for index, employee in enumerate(directors):
        evaluation = evaluations.get(employee.id)
        facility = resolve_facility_for_assignment(employee.assignment)
        facility_key = facility.key if facility else None
        facility_label = facility.label if facility else employee.assignment

        self_score = None
        eval1_score = None
        eval2_score = None
        final_score = None
        low_self_count = None
        low_other_count = None
        salary_raise = None
        rank_order = None
        rank_grade = None
        note = ""
        row_number = 1_000_000 + index

        if facility and facility.bonus_enabled and wb is not None:
            preset = _get_preset(facility.key, wb)
            data_sheet = preset.get("data_sheet")
            if not data_sheet or data_sheet not in wb.sheetnames:
                continue
            ws = wb[preset["data_sheet"]]
            name_row_map = _build_name_row_map(ws, preset)
            excel_row = _resolve_excel_row(employee, name_row_map, aliases)
            if excel_row:
                row_number = excel_row
                self_score = _cell_int(ws.cell(excel_row, preset["self_col"]).value)
                eval1_score = _cell_int(ws.cell(excel_row, preset["eval1_col"]).value)
                eval2_score = _cell_int(ws.cell(excel_row, preset["eval2_col"]).value)
                final_score = _cell_int(ws.cell(excel_row, preset["final_col"]).value)
                rank_order = _cell_int(ws.cell(excel_row, preset["rank_order_col"]).value)
                rank_grade = _cell_str(ws.cell(excel_row, preset["rank_grade_col"]).value) or None
                note = _cell_str(ws.cell(excel_row, preset["note_col"]).value)

        if evaluation:
            totals = _extract_evaluation_totals(employee, evaluation)
            if totals["self_total"] is not None:
                self_score = totals["self_total"]
            if totals["use_eval1"] and totals["eval1_total"] is not None:
                eval1_score = totals["eval1_total"]
            elif not totals["use_eval1"]:
                eval1_score = None
            if totals["eval2_total"] is not None:
                eval2_score = totals["eval2_total"]
            if totals["note"]:
                note = totals["note"]
            final_score = totals["final_score"]

        if final_score is None:
            final_score = _compute_final_score(self_score, eval1_score, eval2_score)

        rows.append(
            _row_dict_from_sources(
                row_number=row_number,
                employee=employee,
                evaluation=evaluation,
                self_score=self_score,
                eval1_score=eval1_score,
                eval2_score=eval2_score,
                final_score=final_score,
                low_self_count=low_self_count,
                low_other_count=low_other_count,
                salary_raise=salary_raise,
                rank_order=rank_order,
                rank_grade=rank_grade,
                note=note,
                bonus_facility_key=facility_key,
                facility_label=facility_label,
            )
        )

    return rows, FACILITY_DIRECTORS_BONUS_LABEL


def _load_bonus_aliases() -> dict[str, str]:
    raw = settings.bonus_name_aliases.strip()
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValueError(f"BONUS_NAME_ALIASES の JSON が不正です: {exc}") from exc
    if not isinstance(parsed, dict):
        raise ValueError("BONUS_NAME_ALIASES は JSON オブジェクトである必要があります")
    return {str(k): str(v) for k, v in parsed.items()}


def _decrypt_workbook(path: Path) -> BytesIO:
    if not path.exists():
        raise FileNotFoundError(f"賞与 Excel が見つかりません: {path}")

    raw = path.read_bytes()
    decrypted = BytesIO()
    office = msoffcrypto.OfficeFile(BytesIO(raw))
    if office.is_encrypted():
        password = settings.bonus_workbook_password.strip()
        if not password:
            raise ValueError("賞与 Excel はパスワード保護されています。BONUS_WORKBOOK_PASSWORD を設定してください")
        office.load_key(password=password)
        office.decrypt(decrypted)
    else:
        decrypted.write(raw)
    decrypted.seek(0)
    return decrypted


def _build_name_row_map(ws, preset: dict[str, Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    name_col = preset["name_col"]
    for row in range(preset["first_data_row"], ws.max_row + 1):
        raw_name = ws.cell(row, name_col).value
        if not raw_name or str(raw_name).strip() in {"職員氏名", ""}:
            continue
        mapping[normalize_person_name(str(raw_name))] = row
    return mapping


def _resolve_excel_row(
    employee: Employee,
    name_row_map: dict[str, int],
    aliases: dict[str, str],
) -> Optional[int]:
    alias_name = aliases.get(employee.employee_id)
    if alias_name:
        row = name_row_map.get(normalize_person_name(alias_name))
        if row:
            return row

    for candidate in (employee.name, format_bonus_name(employee.name)):
        row = name_row_map.get(normalize_person_name(candidate))
        if row:
            return row

    return None


def _apply_row_values(
    ws,
    row: int,
    preset: dict[str, Any],
    *,
    self_total: Optional[int],
    eval1_total: Optional[int],
    eval2_total: Optional[int],
    use_eval1: bool,
    self_scores: dict,
    other_scores: dict,
    note: str,
    apply_cut_rules: bool = True,
) -> None:
    ws.cell(row, preset["self_col"]).value = self_total
    ws.cell(row, preset["eval1_col"]).value = eval1_total if use_eval1 else None
    ws.cell(row, preset["eval2_col"]).value = eval2_total

    if use_eval1 and eval1_total is not None and eval2_total is not None:
        final_formula = f"=({_col_letter(preset['eval1_col'])}{row}+{_col_letter(preset['eval2_col'])}{row})/2"
        ws.cell(row, preset["final_col"]).value = final_formula
    elif eval2_total is not None:
        ws.cell(row, preset["final_col"]).value = f"={_col_letter(preset['eval2_col'])}{row}"
    else:
        ws.cell(row, preset["final_col"]).value = None

    if apply_cut_rules:
        ws.cell(row, preset["low_self_col"]).value = _count_low_self_scores(self_scores)
        ws.cell(row, preset["low_other_col"]).value = _count_low_other_scores(other_scores)
        ws.cell(row, preset["salary_raise_col"]).value = _salary_raise_mark(other_scores)
    else:
        ws.cell(row, preset["low_self_col"]).value = None
        ws.cell(row, preset["low_other_col"]).value = None
        ws.cell(row, preset["salary_raise_col"]).value = None
    ws.cell(row, preset["note_col"]).value = note or None


def _col_letter(col: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(col)


def _template_path() -> Path:
    path = settings.bonus_workbook_template_path.strip()
    if not path:
        raise ValueError("BONUS_WORKBOOK_TEMPLATE_PATH が未設定です")
    return Path(path)


def _open_workbook(*, data_only: bool = False):
    template_path = _template_path()
    decrypted = _decrypt_workbook(template_path)
    return load_workbook(decrypted, data_only=data_only)


def _try_open_workbook(*, data_only: bool = False):
    """Excel テンプレートが無いデモ環境では None を返す。"""
    if not settings.bonus_workbook_template_path.strip():
        return None
    try:
        return _open_workbook(data_only=data_only)
    except FileNotFoundError:
        return None


def _save_workbook(wb) -> None:
    template_path = _template_path()
    backup_path = template_path.with_suffix(template_path.suffix + ".bak")
    if template_path.exists():
        backup_path.write_bytes(template_path.read_bytes())
    buffer = BytesIO()
    wb.save(buffer)
    template_path.write_bytes(buffer.getvalue())


def _cell_int(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _compute_final_score(
    self_score: Optional[int],
    eval1_score: Optional[int],
    eval2_score: Optional[int],
) -> Optional[int]:
    if eval1_score is not None and eval2_score is not None:
        return round((eval1_score + eval2_score) / 2)
    if eval2_score is not None:
        return eval2_score
    if eval1_score is not None:
        return eval1_score
    return self_score


def _set_final_formula(ws, row: int, preset: dict[str, Any], eval1_score, eval2_score) -> None:
    final_col = preset["final_col"]
    if eval1_score is not None and eval2_score is not None:
        ws.cell(row, final_col).value = (
            f"=({_col_letter(preset['eval1_col'])}{row}+{_col_letter(preset['eval2_col'])}{row})/2"
        )
    elif eval2_score is not None:
        ws.cell(row, final_col).value = f"={_col_letter(preset['eval2_col'])}{row}"
    elif eval1_score is not None:
        ws.cell(row, final_col).value = f"={_col_letter(preset['eval1_col'])}{row}"
    else:
        ws.cell(row, final_col).value = None


def _active_period_evaluations(db: Session) -> dict[Any, Evaluation]:
    period = (
        db.query(EvaluationPeriod)
        .filter(EvaluationPeriod.status == PeriodStatus.ACTIVE)
        .first()
    )
    if not period:
        return {}
    evaluations = db.query(Evaluation).filter(Evaluation.period_id == period.id).all()
    return {evaluation.employee_id: evaluation for evaluation in evaluations}


def load_bonus_workbook_rows(
    db: Session,
    *,
    facility_key: str = "inaha",
) -> tuple[list[dict[str, Any]], str]:
    if is_facility_directors_bonus_key(facility_key):
        return _load_directors_bonus_rows(db)

    facility = get_enabled_facility(facility_key)
    wb = _try_open_workbook(data_only=True)

    preset = _get_preset(facility_key, wb)
    data_sheet = preset.get("data_sheet")

    if wb is not None and data_sheet and data_sheet in wb.sheetnames:
        return _load_bonus_rows_from_excel(db, facility_key, preset, wb)

    if not _collect_facility_employees(db, preset):
        return [], facility.label
    return _load_bonus_rows_from_roster(db, facility_key, preset)


def _load_bonus_rows_from_excel(
    db: Session,
    facility_key: str,
    preset: dict[str, Any],
    wb: Any,
) -> tuple[list[dict[str, Any]], str]:
    _ = facility_key
    ws = wb[preset["data_sheet"]]

    aliases = _load_bonus_aliases()
    employees = (
        db.query(Employee)
        .filter(
            Employee.employment_status == EmploymentStatus.ACTIVE,
            Employee.assignment.contains(preset["assignment_contains"]),
        )
        .all()
    )
    employee_by_name: dict[str, Employee] = {}
    for emp in employees:
        employee_by_name[normalize_person_name(emp.name)] = emp
        employee_by_name[normalize_person_name(format_bonus_name(emp.name))] = emp
    for employee_id, alias_name in aliases.items():
        emp = next((e for e in employees if e.employee_id == employee_id), None)
        if emp:
            employee_by_name[normalize_person_name(alias_name)] = emp

    evaluations_by_employee = _active_period_evaluations(db)

    rows: list[dict[str, Any]] = []
    for row in range(preset["first_data_row"], ws.max_row + 1):
        name = _cell_str(ws.cell(row, preset["name_col"]).value)
        if not name or name == "職員氏名":
            continue
        job_title = _cell_str(ws.cell(row, preset["job_col"]).value)
        if job_title.startswith("【") or job_title.startswith("＜"):
            continue

        matched = employee_by_name.get(normalize_person_name(name))
        display_name = format_bonus_name(matched.name) if matched else name
        display_job = bonus_job_title_label(matched) if matched else job_title
        self_score = _cell_int(ws.cell(row, preset["self_col"]).value)
        eval1_score = _cell_int(ws.cell(row, preset["eval1_col"]).value)
        eval2_score = _cell_int(ws.cell(row, preset["eval2_col"]).value)
        final_score = _cell_int(ws.cell(row, preset["final_col"]).value)
        if final_score is None:
            final_score = _compute_final_score(self_score, eval1_score, eval2_score)

        row_data: dict[str, Any] = {
                "row_number": row,
                "employee_id": matched.employee_id if matched else None,
                "name": display_name,
                "job_title": display_job,
                "self_score": self_score,
                "eval1_score": eval1_score,
                "eval2_score": eval2_score,
                "final_score": final_score,
                "low_self_count": _cell_int(ws.cell(row, preset["low_self_col"]).value),
                "low_other_count": _cell_int(ws.cell(row, preset["low_other_col"]).value),
                "salary_raise": _cell_str(ws.cell(row, preset["salary_raise_col"]).value) or None,
                "rank_order": _cell_int(ws.cell(row, preset["rank_order_col"]).value),
                "rank_grade": _cell_str(ws.cell(row, preset["rank_grade_col"]).value) or None,
                "note": _cell_str(ws.cell(row, preset["note_col"]).value),
            }
        if matched:
            row_data.update(
                build_evaluation_insights(
                    matched,
                    evaluations_by_employee.get(matched.id),
                )
            )
            if is_facility_director(matched):
                row_data["low_self_count"] = None
                row_data["low_other_count"] = None
                row_data["salary_raise"] = None
        rows.append(row_data)

    return rows, preset["label"]


def save_bonus_workbook_rows(
    rows: list[dict[str, Any]],
    *,
    facility_key: str = "inaha",
    fiscal_year: Optional[int] = None,
) -> None:
    year = fiscal_year or default_fiscal_year()
    if is_facility_directors_bonus_key(facility_key):
        wb = _open_workbook(data_only=False)
        for item in rows:
            target_key = item.get("bonus_facility_key")
            row = int(item["row_number"])
            if not target_key or row >= 1_000_000:
                continue
            preset = _get_preset(str(target_key), wb)
            data_sheet = preset.get("data_sheet")
            if not data_sheet or data_sheet not in wb.sheetnames:
                continue
            ws = wb[preset["data_sheet"]]
            eval1_score = item.get("eval1_score")
            eval2_score = item.get("eval2_score")
            ws.cell(row, preset["self_col"]).value = item.get("self_score")
            ws.cell(row, preset["eval1_col"]).value = eval1_score
            ws.cell(row, preset["eval2_col"]).value = eval2_score
            _set_final_formula(ws, row, preset, eval1_score, eval2_score)
            ws.cell(row, preset["low_self_col"]).value = item.get("low_self_count")
            ws.cell(row, preset["low_other_col"]).value = item.get("low_other_count")
            ws.cell(row, preset["salary_raise_col"]).value = item.get("salary_raise")
            ws.cell(row, preset["rank_order_col"]).value = item.get("rank_order")
            ws.cell(row, preset["rank_grade_col"]).value = item.get("rank_grade")
            note = item.get("note")
            ws.cell(row, preset["note_col"]).value = note if note else None
        _save_workbook(wb)
        return

    preset = _get_preset(facility_key)
    data_sheet = preset.get("data_sheet")
    wb = None
    try:
        wb = _open_workbook(data_only=False)
    except FileNotFoundError:
        wb = None

    if wb is not None and data_sheet and data_sheet in wb.sheetnames:
        ws = wb[data_sheet]
        for item in rows:
            row = int(item["row_number"])
            eval1_score = item.get("eval1_score")
            eval2_score = item.get("eval2_score")
            ws.cell(row, preset["self_col"]).value = item.get("self_score")
            ws.cell(row, preset["eval1_col"]).value = eval1_score
            ws.cell(row, preset["eval2_col"]).value = eval2_score
            _set_final_formula(ws, row, preset, eval1_score, eval2_score)
            ws.cell(row, preset["low_self_col"]).value = item.get("low_self_count")
            ws.cell(row, preset["low_other_col"]).value = item.get("low_other_count")
            ws.cell(row, preset["salary_raise_col"]).value = item.get("salary_raise")
            ws.cell(row, preset["rank_order_col"]).value = item.get("rank_order")
            ws.cell(row, preset["rank_grade_col"]).value = item.get("rank_grade")
            note = item.get("note")
            ws.cell(row, preset["note_col"]).value = note if note else None
        _save_workbook(wb)

    save_workbook_row_fields(facility_key, year, rows)


def export_bonus_workbook_bytes(*, facility_key: str = "inaha") -> tuple[bytes, str]:
    template_path = _template_path()
    suffix = "_施設長" if is_facility_directors_bonus_key(facility_key) else "_考課反映"
    return template_path.read_bytes(), f"{template_path.stem}{suffix}.xlsx"


def user_can_access_bonus_workbook(user, employee: Optional[Employee]) -> bool:
    if user_has_global_facility_access(user):
        return True
    return bool(employee and is_facility_director(employee))


def _collect_facility_employees(db: Session, preset: dict[str, Any]) -> list[Employee]:
    return (
        db.query(Employee)
        .filter(
            Employee.employment_status == EmploymentStatus.ACTIVE,
            Employee.assignment.contains(preset["assignment_contains"]),
        )
        .all()
    )


def _sync_directors_roster(db: Session, *, dry_run: bool) -> BonusReflectResult:
    wb = _open_workbook(data_only=False)
    aliases = _load_bonus_aliases()
    result = BonusReflectResult(facility=FACILITY_DIRECTORS_BONUS_LABEL)

    for director in _collect_all_facility_directors(db):
        facility = resolve_facility_for_assignment(director.assignment)
        if not facility or not facility.bonus_enabled:
            result.warnings.append(f"{director.name}: 所属施設の賞与シートが未設定です")
            continue
        preset = _get_preset(facility.key, wb)
        data_sheet = preset.get("data_sheet")
        if not data_sheet or data_sheet not in wb.sheetnames:
            result.warnings.append(f"{director.name}: 所属施設の賞与 Excel シートが未設定です")
            continue
        ws = wb[data_sheet]
        name_row_map = _build_name_row_map(ws, preset)
        row = _resolve_excel_row(director, name_row_map, aliases)
        if not row:
            result.unmatched_employees.append(f"{director.employee_id} {director.name}")
            continue
        if not dry_run:
            _apply_employee_identity(ws, row, preset, director)
        result.updated_rows += 1
        result.matched_employees.append(director.name)

    if not dry_run:
        _save_workbook(wb)
    return result


def _reflect_directors_evaluations(
    db: Session,
    period: EvaluationPeriod,
    *,
    dry_run: bool,
) -> tuple[None, BonusReflectResult]:
    template_path = Path(settings.bonus_workbook_template_path.strip())
    aliases = _load_bonus_aliases()
    decrypted = _decrypt_workbook(template_path)
    wb = load_workbook(decrypted)
    evaluations = {
        ev.employee_id: ev
        for ev in db.query(Evaluation).filter(Evaluation.period_id == period.id).all()
    }
    result = BonusReflectResult(facility=FACILITY_DIRECTORS_BONUS_LABEL)
    row_final_scores_by_sheet: dict[str, list[tuple[Any, dict[str, Any], int, int]]] = {}

    for director in _collect_all_facility_directors(db):
        facility = resolve_facility_for_assignment(director.assignment)
        if not facility or not facility.bonus_enabled:
            result.warnings.append(f"{director.name}: 所属施設の賞与シートが未設定です")
            continue
        preset = _get_preset(facility.key, wb)
        data_sheet = preset.get("data_sheet")
        if not data_sheet or data_sheet not in wb.sheetnames:
            result.warnings.append(f"{director.name}: シート「{data_sheet or '未設定'}」が見つかりません")
            continue
        ws = wb[data_sheet]
        name_row_map = _build_name_row_map(ws, preset)
        row = _resolve_excel_row(director, name_row_map, aliases)
        if not row:
            result.unmatched_employees.append(f"{director.employee_id} {director.name}")
            continue

        evaluation = evaluations.get(director.id)
        final_score = _apply_evaluation_for_employee(
            ws,
            row,
            preset,
            director,
            evaluation,
            result,
            dry_run=dry_run,
        )
        if final_score is not None:
            row_final_scores_by_sheet.setdefault(preset["data_sheet"], []).append(
                (ws, preset, row, final_score)
            )
        result.updated_rows += 1
        result.matched_employees.append(director.name)

    if not dry_run and settings.bonus_auto_rank_order:
        for entries in row_final_scores_by_sheet.values():
            ws, preset, _, _ = entries[0]
            ranked = sorted(entries, key=lambda item: (-item[3], item[2]))
            _apply_auto_rank_orders(ws, preset, [(row, score) for _, _, row, score in ranked])

    if not dry_run:
        _save_workbook(wb)
    return None, result


def sync_roster_to_bonus_workbook(
    db: Session,
    *,
    facility_key: str = "inaha",
    dry_run: bool = False,
    fiscal_year: Optional[int] = None,
) -> BonusReflectResult:
    year = fiscal_year or default_fiscal_year()
    if is_facility_directors_bonus_key(facility_key):
        return _sync_directors_roster(db, dry_run=dry_run)

    preset = _get_preset(facility_key)
    employees = _collect_facility_employees(db, preset)
    result = BonusReflectResult(facility=preset["label"])

    wb = None
    try:
        wb = _open_workbook(data_only=False)
    except FileNotFoundError:
        wb = None

    data_sheet = preset.get("data_sheet")
    if wb is None or not data_sheet or data_sheet not in wb.sheetnames:
        result.updated_rows = len(employees)
        result.matched_employees = [employee.name for employee in employees]
        if not dry_run:
            rows, _ = _load_bonus_rows_from_roster(db, facility_key, preset)
            save_workbook_row_fields(facility_key, year, rows)
        return result

    ws = wb[data_sheet]
    name_row_map = _build_name_row_map(ws, preset)
    aliases = _load_bonus_aliases()
    used_rows: set[int] = set()

    for employee in employees:
        row = _resolve_excel_row(employee, name_row_map, aliases)
        if not row:
            result.unmatched_employees.append(f"{employee.employee_id} {employee.name}")
            continue
        if not dry_run:
            _apply_employee_identity(ws, row, preset, employee)
        used_rows.add(row)
        result.updated_rows += 1
        result.matched_employees.append(employee.name)

    for normalized, row in name_row_map.items():
        if row not in used_rows:
            raw = ws.cell(row, preset["name_col"]).value
            if raw:
                result.unmatched_excel_names.append(str(raw))

    if not dry_run:
        _save_workbook(wb)
        rows, _ = _load_bonus_rows_from_roster(db, facility_key, preset)
        save_workbook_row_fields(facility_key, year, rows)
    return result


def reflect_evaluations_to_bonus_workbook(
    db: Session,
    period: EvaluationPeriod,
    *,
    facility_key: str = "inaha",
    dry_run: bool = False,
) -> tuple[Optional[bytes], BonusReflectResult]:
    if is_facility_directors_bonus_key(facility_key):
        return _reflect_directors_evaluations(db, period, dry_run=dry_run)

    preset = _get_preset(facility_key)
    template_path = Path(settings.bonus_workbook_template_path.strip())
    aliases = _load_bonus_aliases()

    wb = None
    if template_path.exists():
        decrypted = _decrypt_workbook(template_path)
        wb = load_workbook(decrypted)

    data_sheet = preset.get("data_sheet")
    if wb is None or not data_sheet or data_sheet not in wb.sheetnames:
        rows, _ = _load_bonus_rows_from_roster(db, facility_key, preset)
        evaluations = {
            ev.employee_id: ev
            for ev in db.query(Evaluation).filter(Evaluation.period_id == period.id).all()
        }
        result = BonusReflectResult(facility=preset["label"])
        row_final_scores: list[tuple[int, int]] = []

        for row_data in rows:
            employee_id = row_data.get("employee_id")
            if not employee_id:
                continue
            employee = next(
                (emp for emp in _collect_facility_employees(db, preset) if emp.employee_id == employee_id),
                None,
            )
            if not employee:
                continue
            evaluation = evaluations.get(employee.id)
            if not evaluation:
                result.warnings.append(f"{employee.name}: 考課データなし")
                continue
            totals = _extract_evaluation_totals(employee, evaluation)
            if not totals["has_scores"]:
                result.warnings.append(f"{employee.name}: 提出済み考課なし")
                continue
            row_data["self_score"] = totals["self_total"]
            row_data["eval1_score"] = totals["eval1_total"] if totals["use_eval1"] else None
            row_data["eval2_score"] = totals["eval2_total"]
            row_data["final_score"] = totals["final_score"]
            row_data["note"] = totals["note"] or row_data.get("note") or ""
            row_data.update(build_evaluation_insights(employee, evaluation))
            if is_facility_director(employee):
                row_data["low_self_count"] = None
                row_data["low_other_count"] = None
                row_data["salary_raise"] = None
            if totals["final_score"] is not None:
                row_final_scores.append((int(row_data["row_number"]), totals["final_score"]))
            result.updated_rows += 1
            result.matched_employees.append(employee.name)

        if settings.bonus_auto_rank_order and row_final_scores:
            ranked = sorted(row_final_scores, key=lambda item: (-item[1], item[0]))
            total = len(ranked)
            order_by_row = {row: order for order, (row, _) in enumerate(ranked, start=1)}
            for row_data in rows:
                row_number = int(row_data["row_number"])
                if row_number not in order_by_row:
                    continue
                order = order_by_row[row_number]
                row_data["rank_order"] = order
                if settings.bonus_auto_rank_grade:
                    row_data["rank_grade"] = grade_for_rank(order, total)

        if dry_run:
            return None, result
        save_workbook_row_fields(facility_key, period.fiscal_year, rows)
        return None, result

    ws = wb[data_sheet]
    name_row_map = _build_name_row_map(ws, preset)
    used_rows: set[int] = set()
    row_final_scores: list[tuple[int, int]] = []

    employees = _collect_facility_employees(db, preset)
    evaluations = {
        ev.employee_id: ev
        for ev in db.query(Evaluation).filter(Evaluation.period_id == period.id).all()
    }

    result = BonusReflectResult(facility=preset["label"])

    for employee in employees:
        row = _resolve_excel_row(employee, name_row_map, aliases)
        if not row:
            result.unmatched_employees.append(f"{employee.employee_id} {employee.name}")
            continue

        evaluation = evaluations.get(employee.id)
        final_score = _apply_evaluation_for_employee(
            ws,
            row,
            preset,
            employee,
            evaluation,
            result,
            dry_run=dry_run,
        )
        if final_score is not None:
            row_final_scores.append((row, final_score))

        used_rows.add(row)
        result.updated_rows += 1
        result.matched_employees.append(employee.name)

    for normalized, row in name_row_map.items():
        if row not in used_rows:
            raw = ws.cell(row, preset["name_col"]).value
            if raw:
                result.unmatched_excel_names.append(str(raw))

    if dry_run:
        return None, result

    if settings.bonus_auto_rank_order and row_final_scores:
        _apply_auto_rank_orders(ws, preset, row_final_scores)
    _save_workbook(wb)
    return template_path.read_bytes(), result
